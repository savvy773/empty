"""
diff.xlsx (또는 지정 파일)에 heis_key / scada_key / match 컬럼을 추가하는 스크립트.

추출 규칙
  heis  : [prefix]/SYSTEM/PATH  →  PATH  (예: SS/VAL_ION8650B_F11M1_W3)
  scada : histprov:.../tag:system/PATH  →  PATH  (예: lgl201/estation/1101a/val_max_source_capability)

비교 규칙 (match 컬럼 E열)
  - Unicode NFKC 정규화
  - 공백·기호(/ _ . : ;) 제거 후 소문자 비교
  - heis_key 기준으로 scada_key 전체 셋에 해당 값이 있으면 MATCH, 없으면 MISSING

사용법
  python add_keys.py                     # 동일 폴더의 diff.xlsx 처리
  python add_keys.py other_file.xlsx     # 파일 직접 지정 (절대/상대 경로 모두 가능)
"""

import os
import re
import sys
import unicodedata
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FILE = os.path.join(SCRIPT_DIR, "diff.xlsx")

HEIS_COL  = "heis"
SCADA_COL = "scada"


def extract_heis_key(value: str) -> str:
    """[prefix]/SYSTEM/ 이후 경로 추출."""
    if not isinstance(value, str) or not value.strip():
        return ""
    m = re.match(r"^\[[^\]]+\]/[^/]+/(.+)$", value.strip())
    return m.group(1) if m else value.strip()


def extract_scada_key(value: str) -> str:
    """/tag:system/ 이후 경로 추출."""
    if not isinstance(value, str) or not value.strip():
        return ""
    m = re.search(r"/tag:[^/]+/(.+)$", value.strip())
    return m.group(1) if m else value.strip()


# 비교용 정규화: NFKC + 기호·공백 제거 + 소문자
_STRIP_PATTERN = re.compile(r"[\s/_.;:]+")

def normalize(value: str) -> str:
    if not isinstance(value, str):
        return ""
    v = unicodedata.normalize("NFKC", value)
    v = _STRIP_PATTERN.sub("", v)
    return v.lower()


def process(filepath: str) -> None:
    if not os.path.exists(filepath):
        print(f"파일 없음: {filepath}")
        sys.exit(1)

    print(f"읽는 중: {filepath}")
    df = pd.read_excel(filepath, dtype=str)

    # 필수 컬럼 확인
    missing = [c for c in (HEIS_COL, SCADA_COL) if c not in df.columns]
    if missing:
        print(f"[오류] 필수 컬럼 없음: {missing}")
        print(f"  실제 컬럼: {list(df.columns)}")
        sys.exit(1)

    df["heis_key"]  = df[HEIS_COL].apply(extract_heis_key)
    df["scada_key"] = df[SCADA_COL].apply(extract_scada_key)

    # 양방향 정규화 셋 구성
    scada_norm_set = {normalize(v) for v in df["scada_key"] if v}
    heis_norm_set  = {normalize(v) for v in df["heis_key"]  if v}

    def check_match(row) -> str:
        hk = row["heis_key"]
        sk = row["scada_key"]
        # MISSING: scada_key 자체가 없음 (매핑 미제공)
        if not sk:
            return "MISSING"
        # MATCH: 양방향 정규화 비교에서 하나라도 일치
        if (hk and normalize(hk) in scada_norm_set) or \
           (normalize(sk) in heis_norm_set):
            return "MATCH"
        # UNMATCH: scada_key 있지만 어느 쪽 셋에도 매칭 없음
        return "UNMATCH"

    df["match"] = df.apply(check_match, axis=1)
    # scada_key 는 B열 추출값 그대로 유지 (교체 없음)

    # 정렬: MISSING → UNMATCH → MATCH
    sort_order = {"MISSING": 0, "UNMATCH": 1, "MATCH": 2, "": 3}
    df["_sort"] = df["match"].map(lambda v: sort_order.get(v, 3))
    df = df.sort_values("_sort", kind="stable").drop(columns=["_sort"]).reset_index(drop=True)

    missing_count = (df["match"] == "MISSING").sum()
    match_count   = (df["match"] == "MATCH").sum()

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="diff", na_rep="")
        ws = writer.sheets["diff"]
        ws.freeze_panes = "B2"

        # 색상 정의
        blue    = PatternFill("solid", fgColor="4472C4")
        green   = PatternFill("solid", fgColor="70AD47")
        orange  = PatternFill("solid", fgColor="FF6600")
        red_fill     = PatternFill("solid", fgColor="FF4444")
        yellow_fill  = PatternFill("solid", fgColor="FFD966")
        green_fill   = PatternFill("solid", fgColor="92D050")

        # 헤더 스타일
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            if cell.column <= 2:
                cell.fill = blue
            elif cell.column == 5:          # match 컬럼
                cell.fill = orange
            else:
                cell.fill = green

        # match 컬럼 셀 색상 (MATCH=연두, MISSING=빨강)
        match_col_idx = df.columns.get_loc("match") + 1   # 1-based
        for row in ws.iter_rows(min_row=2, min_col=match_col_idx, max_col=match_col_idx):
            for cell in row:
                if cell.value == "MISSING":
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                elif cell.value == "UNMATCH":
                    cell.fill = yellow_fill
                    cell.font = Font(bold=True)
                elif cell.value == "MATCH":
                    cell.fill = green_fill

        # 전체 필터
        ws.auto_filter.ref = ws.dimensions

        # 컬럼 너비
        for col, width in zip("ABCDE", [60, 80, 40, 55, 12]):
            ws.column_dimensions[col].width = width

    unmatch_count = (df["match"] == "UNMATCH").sum()
    print(f"저장 완료: {filepath}")
    print(f"총 {len(df):,} rows  |  MATCH: {match_count}  UNMATCH: {unmatch_count}  MISSING: {missing_count}")

    for label in ("MISSING", "UNMATCH"):
        subset = df[df["match"] == label][["heis_key", "scada_key"]]
        if not subset.empty:
            print(f"\n--- {label} 목록 ({len(subset)}건) ---")
            print(subset.to_string(index=False))


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    # 상대 경로이면 스크립트 위치 기준으로 해석
    if not os.path.isabs(target):
        target = os.path.join(SCRIPT_DIR, target)
    process(target)
